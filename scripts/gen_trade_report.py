"""
生成回测交易报告 Excel，包含三个 sheet：
  1. 交易明细    - 所有买入/卖出记录（去掉持有行）
  2. 配对记录    - 每笔买入与对应卖出配对，显示完整盈亏
  3. 标的汇总    - 每只股票的总交易次数、股数、盈亏统计

运行：python scripts/gen_trade_report.py
"""
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent))

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

INPUT  = Path("回溯交易记录/trades_a_detail.csv")
OUTPUT = Path("回溯交易记录/回测交易报告.xlsx")


def load_data():
    df = pd.read_csv(INPUT, dtype={"代码": str})
    df["日期"] = pd.to_datetime(df["日期"])
    df["股数(股)"]  = pd.to_numeric(df["股数(股)"],  errors="coerce")
    df["成本价(元)"] = pd.to_numeric(df["成本价(元)"], errors="coerce")
    df["成交价(元)"] = pd.to_numeric(df["成交价(元)"], errors="coerce")
    df["成交金额(元)"] = pd.to_numeric(df["成交金额(元)"], errors="coerce")
    df["盈亏(元)"]  = pd.to_numeric(df["盈亏(元)"],  errors="coerce").fillna(0)
    df["盈亏(%)"]   = pd.to_numeric(df["盈亏(%)"],   errors="coerce").fillna(0)
    return df


def make_detail(df: pd.DataFrame) -> pd.DataFrame:
    """Sheet1：买入/卖出明细，去掉持有行，列名更直观。"""
    detail = df[df["方向"].isin(["买入", "卖出"])].copy()
    detail = detail.rename(columns={
        "日期":      "成交日期",
        "成本价(元)": "买入均价(元)",
        "成交价(元)": "成交价(元)",
    })
    cols = ["成交日期","方向","代码","名称","行业",
            "手数(手)","股数(股)","买入均价(元)","成交价(元)",
            "成交金额(元)","手续费(元)","盈亏(元)","盈亏(%)","备注"]
    return detail[cols].reset_index(drop=True)


def make_paired(df: pd.DataFrame) -> pd.DataFrame:
    """Sheet2：配对记录——将每只股票的买入和对应卖出配成一行。"""
    buys  = df[df["方向"] == "买入"].copy()
    sells = df[df["方向"] == "卖出"].copy()

    rows = []
    for code in df["代码"].unique():
        name = df[df["代码"] == code]["名称"].iloc[0] if not df[df["代码"] == code].empty else ""
        ind  = df[df["代码"] == code]["行业"].iloc[0] if not df[df["代码"] == code].empty else ""
        b = buys[buys["代码"] == code].sort_values("日期").reset_index(drop=True)
        s = sells[sells["代码"] == code].sort_values("日期").reset_index(drop=True)

        # 按顺序配对
        n = min(len(b), len(s))
        for i in range(n):
            buy_row  = b.iloc[i]
            sell_row = s.iloc[i]
            shares   = sell_row["股数(股)"]
            buy_p    = buy_row["成交价(元)"]
            sell_p   = sell_row["成交价(元)"]
            pnl      = sell_row["盈亏(元)"]
            days     = (sell_row["日期"] - buy_row["日期"]).days if pd.notna(sell_row["日期"]) else 0
            pnl_pct  = round((sell_p / buy_p - 1) * 100, 2) if buy_p and not np.isnan(buy_p) else 0

            rows.append({
                "代码":     code,
                "名称":     name,
                "行业":     ind,
                "买入日期":  buy_row["日期"].strftime("%Y-%m-%d") if pd.notna(buy_row["日期"]) else "",
                "买入价(元)": round(buy_p, 2) if pd.notna(buy_p) else "",
                "卖出日期":  sell_row["日期"].strftime("%Y-%m-%d") if pd.notna(sell_row["日期"]) else "",
                "卖出价(元)": round(sell_p, 2) if pd.notna(sell_p) else "",
                "持仓天数":  days,
                "股数(股)":  int(shares) if pd.notna(shares) else 0,
                "手数(手)":  int(shares // 100) if pd.notna(shares) else 0,
                "盈亏(元)":  round(pnl, 2),
                "盈亏(%)":   pnl_pct,
                "结果":     "盈利" if pnl > 0 else ("亏损" if pnl < 0 else "持平"),
            })
        # 未配对的买入（最后仍持有）
        for i in range(n, len(b)):
            buy_row = b.iloc[i]
            buy_p   = buy_row["成交价(元)"]
            rows.append({
                "代码":     code,
                "名称":     name,
                "行业":     ind,
                "买入日期":  buy_row["日期"].strftime("%Y-%m-%d") if pd.notna(buy_row["日期"]) else "",
                "买入价(元)": round(buy_p, 2) if pd.notna(buy_p) else "",
                "卖出日期":  "（持有中）",
                "卖出价(元)": "",
                "持仓天数":  "",
                "股数(股)":  int(buy_row["股数(股)"]) if pd.notna(buy_row["股数(股)"]) else 0,
                "手数(手)":  int(buy_row["股数(股)"] // 100) if pd.notna(buy_row["股数(股)"]) else 0,
                "盈亏(元)":  "",
                "盈亏(%)":   "",
                "结果":     "持有中",
            })

    paired = pd.DataFrame(rows)
    return paired.sort_values(["代码", "买入日期"]).reset_index(drop=True)


def make_summary(df: pd.DataFrame, paired: pd.DataFrame) -> pd.DataFrame:
    """Sheet3：按标的汇总买卖次数、总股数、总盈亏。"""
    buys  = df[df["方向"] == "买入"]
    sells = df[df["方向"] == "卖出"]

    buy_stats  = buys.groupby("代码").agg(
        买入次数=("日期", "count"),
        总买入股数=("股数(股)", "sum"),
        首次买入日=("日期", "min"),
    )
    sell_stats = sells.groupby("代码").agg(
        卖出次数=("日期", "count"),
        总卖出股数=("股数(股)", "sum"),
        总盈亏=("盈亏(元)", "sum"),
        盈利笔数=("盈亏(元)", lambda x: (x > 0).sum()),
        亏损笔数=("盈亏(元)", lambda x: (x < 0).sum()),
        最大单笔盈亏=("盈亏(元)", lambda x: x.loc[x.abs().idxmax()] if len(x) else 0),
    )

    name_map = df.groupby("代码")["名称"].first()
    ind_map  = df.groupby("代码")["行业"].first()

    summary = buy_stats.join(sell_stats, how="outer").join(name_map).join(ind_map)
    summary.insert(0, "名称", summary.pop("名称"))
    summary.insert(1, "行业", summary.pop("行业"))
    summary["胜率(%)"] = (summary["盈利笔数"] / summary["卖出次数"] * 100).round(1)
    summary["总盈亏"]  = summary["总盈亏"].round(2)
    summary = summary.reset_index()

    summary["总盈亏"] = pd.to_numeric(summary["总盈亏"], errors="coerce")
    summary["胜率(%)"] = pd.to_numeric(summary["胜率(%)"], errors="coerce")
    # 排序：只对有卖出记录的排序，无卖出（仍持有）放最后
    has_sell = summary["卖出次数"].notna() & (summary["卖出次数"] > 0)
    summary = pd.concat([
        summary[has_sell].sort_values("总盈亏", ascending=False),
        summary[~has_sell].sort_values("首次买入日"),
    ]).reset_index(drop=True)
    return summary


def make_field_desc() -> pd.DataFrame:
    """Sheet4：字段说明——每列的含义和取值规则。"""
    rows = [
        # ── 交易明细 ──
        ("交易明细", "成交日期",   "调仓执行日期，格式 YYYY-MM-DD"),
        ("交易明细", "方向",       "买入 / 卖出（持有行已过滤）"),
        ("交易明细", "代码",       "A股6位股票代码"),
        ("交易明细", "名称",       "股票简称"),
        ("交易明细", "行业",       "申万一级行业分类"),
        ("交易明细", "手数(手)",   "成交手数，1手=100股；科创板最小2手=200股"),
        ("交易明细", "股数(股)",   "成交股数=手数×100（科创板×200）"),
        ("交易明细", "买入均价(元)","买入行=成交价；卖出行=该笔持仓买入时的成本均价"),
        ("交易明细", "成交价(元)", "本次实际成交参考价（回测用调仓日收盘价）"),
        ("交易明细", "成交金额(元)","成交价 × 股数"),
        ("交易明细", "手续费(元)", "佣金按0.125%单边计算（含印花税）"),
        ("交易明细", "盈亏(元)",   "买入行=0；卖出行=(卖出价−成本价)×股数−手续费"),
        ("交易明细", "盈亏(%)",    "买入行=0；卖出行=(卖出价÷成本价−1)×100"),
        ("交易明细", "备注",       "纳入持仓池/移出持仓池/资金不足一手（说明未能建仓）"),
        # ── 配对记录 ──
        ("配对记录", "代码/名称/行业", "同交易明细"),
        ("配对记录", "买入日期",   "该笔仓位的买入日期"),
        ("配对记录", "买入价(元)", "买入时的参考收盘价（即成本价）"),
        ("配对记录", "卖出日期",   "卖出日期；'（持有中）'表示回测截止时仍持有"),
        ("配对记录", "卖出价(元)", "卖出时的参考收盘价"),
        ("配对记录", "持仓天数",   "从买入到卖出的自然日数"),
        ("配对记录", "股数(股)",   "该笔交易的股数"),
        ("配对记录", "手数(手)",   "该笔交易的手数"),
        ("配对记录", "盈亏(元)",   "本笔已实现盈亏；持有中=空"),
        ("配对记录", "盈亏(%)",    "本笔已实现盈亏百分比；持有中=空"),
        ("配对记录", "结果",       "盈利 / 亏损 / 持平 / 持有中"),
        # ── 标的汇总 ──
        ("标的汇总", "代码/名称/行业", "同交易明细"),
        ("标的汇总", "买入次数",   "该股票在回测期内的买入次数"),
        ("标的汇总", "总买入股数", "历次买入股数合计"),
        ("标的汇总", "首次买入日", "策略第一次买入该股票的日期"),
        ("标的汇总", "卖出次数",   "历次卖出次数（可能<买入次数，差值=当前仍持有次数）"),
        ("标的汇总", "总卖出股数", "历次卖出股数合计"),
        ("标的汇总", "总盈亏",     "所有已实现盈亏之和（元），正=盈，负=亏"),
        ("标的汇总", "盈利笔数",   "卖出盈利的次数"),
        ("标的汇总", "亏损笔数",   "卖出亏损的次数"),
        ("标的汇总", "最大单笔盈亏","绝对值最大的单笔盈亏（元）"),
        ("标的汇总", "胜率(%)",    "盈利笔数 ÷ 卖出次数 × 100"),
        # ── 说明 ──
        ("全局说明", "固定等权分配", "每笔建仓目标金额=60万÷30=2万元；高价股若2万不足一手则标注'资金不足'"),
        ("全局说明", "与回测收益的关系", "本报告用固定2万/只模拟，不随组合净值增长；实际回测用百分比收益，两者盈亏数字不同"),
        ("全局说明", "科创板",     "代码688开头，最小申报单位200股（2手），非100股"),
        ("全局说明", "持仓成本",   "取买入当日参考价（调仓日收盘价），未考虑多次加仓均价"),
    ]
    return pd.DataFrame(rows, columns=["所属Sheet", "字段名", "说明"])


def style_sheet(ws, header_fill="#1F3A5F", alt_fill="#EEF2F7"):
    """给 sheet 加样式：表头蓝底白字，数据行隔行底纹，数值右对齐。"""
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    hdr_fill = PatternFill("solid", fgColor=header_fill.lstrip("#"))
    alt      = PatternFill("solid", fgColor=alt_fill.lstrip("#"))
    thin     = Side(style="thin", color="CCCCCC")
    border   = Border(left=thin, right=thin, top=thin, bottom=thin)
    center   = Alignment(horizontal="center", vertical="center")
    right    = Alignment(horizontal="right",  vertical="center")

    for col_idx, cell in enumerate(ws[1], 1):
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = center
        cell.border    = border

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
        bg = alt if row_idx % 2 == 0 else None
        for cell in row:
            if bg:
                cell.fill = bg
            cell.border = border
            if isinstance(cell.value, float):
                cell.alignment = right
            elif isinstance(cell.value, int):
                cell.alignment = right

    # 自动列宽
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val_len = len(str(cell.value)) if cell.value is not None else 0
                # 中文字符宽度约为2
                cn_count = sum(1 for c in str(cell.value or "") if "一" <= c <= "鿿")
                val_len = val_len + cn_count
                max_len = max(max_len, val_len)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 25)


def color_pnl(ws, pnl_col_idx: int):
    """盈亏列：正数标绿，负数标红。"""
    red   = Font(color="CC0000", bold=True)
    green = Font(color="007A29", bold=True)
    for row in ws.iter_rows(min_row=2, min_col=pnl_col_idx, max_col=pnl_col_idx):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                if cell.value > 0:
                    cell.font = green
                elif cell.value < 0:
                    cell.font = red


def main():
    print("加载数据...")
    df = load_data()

    print("生成交易明细...")
    detail = make_detail(df)

    print("生成配对记录...")
    paired = make_paired(df)

    print("生成标的汇总...")
    summary = make_summary(df, paired)

    print("生成字段说明...")
    field_desc = make_field_desc()

    print(f"写入 Excel → {OUTPUT}")
    with pd.ExcelWriter(OUTPUT, engine="openpyxl") as writer:
        field_desc.to_excel(writer, sheet_name="字段说明", index=False)
        detail.to_excel(writer,     sheet_name="交易明细", index=False)
        paired.to_excel(writer,     sheet_name="配对记录", index=False)
        summary.to_excel(writer,    sheet_name="标的汇总", index=False)

    # 美化 + 冻结标题行
    wb = load_workbook(OUTPUT)
    for sname in ["字段说明", "交易明细", "配对记录", "标的汇总"]:
        ws = wb[sname]
        style_sheet(ws)
        ws.freeze_panes = "A2"   # 冻结第一行，滚动时始终可见列名

    # 盈亏列着色
    for sname, col_name in [("交易明细", "盈亏(元)"), ("配对记录", "盈亏(元)"), ("标的汇总", "总盈亏"),
                             ("标的汇总", "最大单笔盈亏")]:
        ws = wb[sname]
        headers = [c.value for c in ws[1]]
        if col_name in headers:
            color_pnl(ws, headers.index(col_name) + 1)

    wb.save(OUTPUT)

    # 控制台摘要
    sell_pnl = df[df["方向"] == "卖出"]["盈亏(元)"]
    print(f"\n=== 报告摘要 ===")
    print(f"交易明细：{len(detail)} 笔（买入 {len(detail[detail['方向']=='买入'])} / 卖出 {len(detail[detail['方向']=='卖出'])}）")
    print(f"配对记录：{len(paired)} 条")
    print(f"涉及标的：{summary['代码'].nunique()} 只")
    print(f"实现盈亏：{sell_pnl.sum():+,.0f} 元")
    print(f"胜率    ：{(sell_pnl > 0).sum()} 盈 / {(sell_pnl < 0).sum()} 亏")
    print(f"\n盈亏 Top5：")
    print(summary[["代码","名称","卖出次数","总盈亏","胜率(%)"]].head(5).to_string(index=False))
    print(f"\n亏损 Top5：")
    print(summary[["代码","名称","卖出次数","总盈亏","胜率(%)"]].tail(5).to_string(index=False))
    print(f"\n已保存 → {OUTPUT}")


if __name__ == "__main__":
    main()
