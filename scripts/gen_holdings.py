"""生成当日完整持仓明细，含买入价、持仓天数、浮动盈亏等。"""
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent))

import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BUILD_DATE = "2026-05-11"
TODAY_DATE = "2026-05-15"
DAYS_HELD  = (datetime.strptime(TODAY_DATE, "%Y-%m-%d") -
              datetime.strptime(BUILD_DATE, "%Y-%m-%d")).days

LOGS_DIR   = Path("logs")
OUT_DIR    = Path("回溯交易记录")


def load_track_a():
    start = pd.read_csv(LOGS_DIR / "paper_trade_a_start.csv", dtype={"代码": str})
    today_file = LOGS_DIR / f"paper_trade_{TODAY_DATE.replace('-','')}.csv"
    if not today_file.exists():
        print(f"今日数据文件不存在：{today_file}")
        return pd.DataFrame()
    today = pd.read_csv(today_file, dtype={"代码": str})
    today = today[today["代码"] != "合计"].copy()
    cur_col = [c for c in today.columns if "收盘" in c][0]
    today = today.rename(columns={cur_col: "今日收盘价(元)"})

    df = today.merge(start[["代码", "备注"]], on="代码", how="left")
    result = pd.DataFrame({
        "代码":         df["代码"],
        "名称":         df["名称"],
        "行业":         df["行业"],
        "建仓日期":     BUILD_DATE,
        "建仓价(元)":   pd.to_numeric(df["建仓价(元)"], errors="coerce"),
        "手数(手)":     (pd.to_numeric(df["股数(股)"], errors="coerce") / 100).astype("Int64"),
        "股数(股)":     pd.to_numeric(df["股数(股)"], errors="coerce"),
        "建仓金额(元)": pd.to_numeric(df["建仓金额(元)"], errors="coerce"),
        "今日收盘价(元)": pd.to_numeric(df["今日收盘价(元)"], errors="coerce"),
        "今日市值(元)": pd.to_numeric(df["当前市值(元)"], errors="coerce"),
        "持仓天数":     DAYS_HELD,
        "浮动盈亏(元)": pd.to_numeric(df["浮动盈亏(元)"], errors="coerce"),
        "涨跌幅(%)":   pd.to_numeric(df["涨跌幅(%)"], errors="coerce"),
        "状态":         df["状态"],
        "备注":         df["备注"].fillna(""),
    })
    return result.sort_values("涨跌幅(%)", ascending=False).reset_index(drop=True)


def load_track_b():
    b_start_file = LOGS_DIR / "paper_trade_b_start.csv"
    b_today_file = LOGS_DIR / f"paper_trade_b_{TODAY_DATE.replace('-','')}.csv"
    if not b_start_file.exists() or not b_today_file.exists():
        return pd.DataFrame()
    b_start = pd.read_csv(b_start_file, dtype={"代码": str})
    b_today = pd.read_csv(b_today_file, dtype={"代码": str})
    return b_start.merge(b_today, on="代码", how="left", suffixes=("_start", "_today"))


def add_summary_row(df: pd.DataFrame) -> pd.DataFrame:
    """在末尾追加合计行。"""
    cost  = df["建仓金额(元)"].sum()
    mktv  = df["今日市值(元)"].sum()
    pnl   = df["浮动盈亏(元)"].sum()
    pct   = round((mktv / cost - 1) * 100, 2) if cost else 0
    row = {c: "" for c in df.columns}
    row.update({
        "代码": "【合计】",
        "手数(手)": df["手数(手)"].sum(),
        "股数(股)": df["股数(股)"].sum(),
        "建仓金额(元)": round(cost, 2),
        "今日市值(元)": round(mktv, 2),
        "持仓天数": DAYS_HELD,
        "浮动盈亏(元)": round(pnl, 2),
        "涨跌幅(%)": pct,
        "状态": "盈利" if pnl >= 0 else "亏损",
    })
    return pd.concat([df, pd.DataFrame([row])], ignore_index=True)


def style_ws(ws):
    hdr_fill = PatternFill("solid", fgColor="1F3A5F")
    alt_fill = PatternFill("solid", fgColor="EEF2F7")
    sum_fill = PatternFill("solid", fgColor="FFF3CD")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    last_row = ws.max_row
    for i, row in enumerate(ws.iter_rows(min_row=2), 2):
        is_summary = ws.cell(i, 1).value == "【合计】"
        for cell in row:
            cell.border = border
            if is_summary:
                cell.fill = sum_fill
                cell.font = Font(bold=True, size=10)
            elif i % 2 == 0:
                cell.fill = alt_fill
            if isinstance(cell.value, (float, int)) and not is_summary:
                cell.alignment = Alignment(horizontal="right")

    # 盈亏列着色
    headers = [c.value for c in ws[1]]
    for col_name in ["浮动盈亏(元)", "涨跌幅(%)"]:
        if col_name in headers:
            ci = headers.index(col_name) + 1
            red, green = Font(color="CC0000", bold=True), Font(color="007A29", bold=True)
            for row in ws.iter_rows(min_row=2, min_col=ci, max_col=ci):
                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        if cell.value > 0:
                            cell.font = green
                        elif cell.value < 0:
                            cell.font = red

    # 自动列宽
    for col in ws.columns:
        max_len = 0
        cl = get_column_letter(col[0].column)
        for cell in col:
            v = str(cell.value or "")
            max_len = max(max_len, len(v) + sum(1 for c in v if "一" <= c <= "鿿"))
        ws.column_dimensions[cl].width = min(max_len + 3, 22)

    ws.freeze_panes = "A2"


def print_summary(df: pd.DataFrame, title: str):
    data = df[df["代码"] != "【合计】"]
    total = df[df["代码"] == "【合计】"].iloc[0]
    print(f"\n{'='*65}")
    print(f"  {title}  持仓明细  {TODAY_DATE}  （建仓日：{BUILD_DATE}，持仓{DAYS_HELD}天）")
    print(f"{'='*65}")
    print(data[["代码","名称","行业","建仓价(元)","今日收盘价(元)","涨跌幅(%)","浮动盈亏(元)","状态"]].to_string(index=False))
    print(f"{'-'*65}")
    print(f"  建仓总额 {float(total['建仓金额(元)']):>12,.0f} 元")
    print(f"  今日市值 {float(total['今日市值(元)']):>12,.0f} 元")
    print(f"  浮动盈亏 {float(total['浮动盈亏(元)']):>+12,.0f} 元  ({float(total['涨跌幅(%)']):+.2f}%)")
    up = len(data[data["涨跌幅(%)"] > 0])
    dn = len(data[data["涨跌幅(%)"] < 0])
    no = len(data) - up - dn
    print(f"  盈利 {up} 只 / 亏损 {dn} 只 / 无数据 {no} 只")


def main():
    track_a = load_track_a()
    if track_a.empty:
        print("Track A 数据加载失败")
        return

    track_a_full = add_summary_row(track_a)
    print_summary(track_a_full, "Track A")

    out_csv   = OUT_DIR / f"持仓明细_{TODAY_DATE}.csv"
    out_excel = OUT_DIR / f"持仓明细_{TODAY_DATE}.xlsx"

    track_a_full.to_csv(out_csv, index=False, encoding="utf-8-sig")

    with pd.ExcelWriter(out_excel, engine="openpyxl") as writer:
        track_a_full.to_excel(writer, sheet_name=f"Track A 持仓 {TODAY_DATE}", index=False)
        # 字段说明
        desc = pd.DataFrame([
            ("代码",         "A股6位股票代码"),
            ("名称",         "股票简称"),
            ("行业",         "申万一级行业分类"),
            ("建仓日期",     f"纸面交易建仓日期：{BUILD_DATE}"),
            ("建仓价(元)",   "建仓参考价（2026-05-30收盘价）"),
            ("手数(手)",     "持仓手数，1手=100股；科创板2手=200股"),
            ("股数(股)",     "持仓总股数"),
            ("建仓金额(元)", "建仓价 × 股数（实际建仓总成本）"),
            ("今日收盘价(元)","今日（2026-05-15）A股收盘价"),
            ("今日市值(元)", "今日收盘价 × 股数"),
            ("持仓天数",     f"自建仓日至今的自然日数（{DAYS_HELD}天）"),
            ("浮动盈亏(元)", "今日市值 − 建仓金额（未实现盈亏）"),
            ("涨跌幅(%)",    "（今日价 ÷ 建仓价 − 1）× 100"),
            ("状态",         "正常 / 无数据（当日停牌或数据缺失）"),
            ("备注",         "超2万(最小1手) = 该股单手成本超2万元，无法等额均摊"),
        ], columns=["字段名", "说明"])
        desc.to_excel(writer, sheet_name="字段说明", index=False)

    wb = load_workbook(out_excel)
    for sname in wb.sheetnames:
        ws = wb[sname]
        style_ws(ws)
    wb.save(out_excel)

    print(f"\n  已保存 CSV   → {out_csv}")
    print(f"  已保存 Excel → {out_excel}")


if __name__ == "__main__":
    main()
